from flask import Flask, request
from openpyxl import load_workbook, Workbook
from datetime import datetime
import requests
import os
import threading
import queue
import time
import re


# ==========================================
# CONFIGURACIÓN
# ==========================================

app = Flask(__name__)

PHONE_NUMBER_ID = "1248995224968397"

IMAGEN_UBICACION = "PIMENTEL.jpg"

IMAGEN_CASA = "CASA PUERTO REAL 2026.jpeg"

TOKEN_VERIFICACION = "puerto_real_2026"

ARCHIVO_EXCEL = "clientes_puerto_real.xlsx"

GOOGLE_SHEETS_URL = "https://script.google.com/macros/s/AKfycbzMSUCODNw21Jsv4cUiBpZosSBIRbHFQThC-JipvhJ1r6ok1So0cde7D4GbuFx6r4Jp/exec"

# ==========================================
# MEMORIA DE CONVERSACIONES
# ==========================================

historiales = {}


# ==========================================
# CONTROL DE DUPLICADOS Y COLA DE TRABAJO
# ==========================================

ARCHIVO_MENSAJES_PROCESADOS = "mensajes_procesados.txt"

mensajes_procesados = set()
lock_mensajes = threading.Lock()
cola_mensajes = queue.Queue()
ultimo_mensaje_por_numero = {}

ultimos_mensajes_por_cliente = {}
VENTANA_DUPLICADO_SEGUNDOS = 12


def cargar_mensajes_procesados():
    if not os.path.exists(ARCHIVO_MENSAJES_PROCESADOS):
        return

    try:
        with open(ARCHIVO_MENSAJES_PROCESADOS, "r", encoding="utf-8") as archivo:
            for linea in archivo:
                mensaje_id = linea.strip()
                if mensaje_id:
                    mensajes_procesados.add(mensaje_id)
    except Exception as error:
        print("No se pudo cargar mensajes procesados:", error)


def registrar_mensaje_procesado(mensaje_id):
    if not mensaje_id:
        return True

    with lock_mensajes:
        if mensaje_id in mensajes_procesados:
            return False

        mensajes_procesados.add(mensaje_id)

        try:
            with open(ARCHIVO_MENSAJES_PROCESADOS, "a", encoding="utf-8") as archivo:
                archivo.write(mensaje_id + "\n")
        except Exception as error:
            print("No se pudo guardar el ID del mensaje:", error)

    return True


# ==========================================
# INFORMACIÓN DE PUERTO REAL
# ==========================================

informacion_puerto_real = """
PROYECTO: PUERTO REAL

Ubicación:
Pimentel, Chiclayo.

Tipo:
Viviendas.

Características:
- 2 pisos.
- 3 habitaciones.
- 2 baños.

Condominio:
- Condominio privado.
- Cerco perimétrico de concreto.
- Incluye 4 parques.

Construcción:
- Sistema constructivo tradicional de ladrillo y cemento.

Áreas:
- Área total: 60 m².
- Área construida: 60 m².
- Primer piso: 33 m².
- Segundo piso: 27 m².

Facilidades económicas:
- Inicial: S/ 5,000.
- Bono de Techo Propio: S/ 62,700.
- Financiamiento: 48 meses sin intereses.

Servicios básicos:
- Luz.
- Agua.
- Desagüe.

Entrega:
- Se entrega en 30 meses.
"""


# ==========================================
# EXCEL
# ==========================================

ENCABEZADOS_EXCEL = [
    "Fecha",
    "Hora",
    "Nombre",
    "WhatsApp",
    "ID WhatsApp",
    "Username",
    "Motivo de compra",
    "Número de personas",
    "Inicial",
    "Bono",
    "Ubicación",
    "Financiamiento",
    "Solicitud de asesor",
    "Horario de contacto",
    "Nivel del lead",
    "Resumen",
]


def preparar_excel():
    if not os.path.exists(ARCHIVO_EXCEL):
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Clientes"
        hoja.append(ENCABEZADOS_EXCEL)
        libro.save(ARCHIVO_EXCEL)
        return

    # Asegura que la fila de encabezados tenga las 16 columnas actuales.
    try:
        libro = load_workbook(ARCHIVO_EXCEL)
        hoja = libro.active
        for columna, encabezado in enumerate(ENCABEZADOS_EXCEL, start=1):
            hoja.cell(row=1, column=columna).value = encabezado
        libro.save(ARCHIVO_EXCEL)
    except PermissionError:
        print("ERROR: CIERRA EL EXCEL ANTES DE INICIAR EL SERVIDOR.")


def dato_valido(valor):
    if valor is None:
        return False

    valor = str(valor).strip()

    if valor == "":
        return False

    if valor.lower() in [
        "no especificado",
        "no especificada",
        "desconocido",
        "desconocida",
    ]:
        return False

    return True


def guardar_cliente_excel(
    nombre,
    whatsapp,
    id_whatsapp,
    username,
    motivo_compra,
    numero_personas,
    inicial,
    bono,
    ubicacion,
    financiamiento,
    solicitud_asesor,
    horario_contacto,
    nivel_lead,
    resumen,
):
    preparar_excel()

    ahora = datetime.now()
    fecha = ahora.strftime("%d/%m/%Y")
    hora = ahora.strftime("%H:%M")

    try:
        libro = load_workbook(ARCHIVO_EXCEL)
    except PermissionError:
        print()
        print("ERROR: CIERRA EL EXCEL ANTES DE GUARDAR.")
        print()
        return

    hoja = libro.active
    fila_encontrada = None

    for fila in range(2, hoja.max_row + 1):
        whatsapp_excel = hoja.cell(row=fila, column=4).value
        id_excel = hoja.cell(row=fila, column=5).value

        if (
            str(id_excel).strip() == str(id_whatsapp).strip()
            and str(id_whatsapp).strip() not in ["", "None", "No especificado"]
        ):
            fila_encontrada = fila
            break

        if (
            str(whatsapp_excel).strip() == str(whatsapp).strip()
            and str(whatsapp).strip() not in ["", "None", "No especificado"]
        ):
            fila_encontrada = fila
            break

    nuevos_datos = {
        3: nombre,
        4: whatsapp,
        5: id_whatsapp,
        6: username,
        7: motivo_compra,
        8: numero_personas,
        9: inicial,
        10: bono,
        11: ubicacion,
        12: financiamiento,
        13: solicitud_asesor,
        14: horario_contacto,
        15: nivel_lead,
        16: resumen,
    }

    if fila_encontrada:
        hoja.cell(row=fila_encontrada, column=1).value = fecha
        hoja.cell(row=fila_encontrada, column=2).value = hora

        for columna, valor in nuevos_datos.items():
            if dato_valido(valor):
                hoja.cell(row=fila_encontrada, column=columna).value = valor

        print("CLIENTE ACTUALIZADO EN EXCEL")

    else:
        hoja.append([
            fecha,
            hora,
            nombre,
            whatsapp,
            id_whatsapp,
            username,
            motivo_compra,
            numero_personas,
            inicial,
            bono,
            ubicacion,
            financiamiento,
            solicitud_asesor,
            horario_contacto,
            nivel_lead,
            resumen,
        ])

        print("CLIENTE GUARDADO EN EXCEL")

    libro.save(ARCHIVO_EXCEL)
def guardar_cliente_google_sheets(
    nombre,
    whatsapp,
    id_whatsapp,
    username,
    motivo_compra,
    numero_personas,
    inicial,
    bono,
    ubicacion,
    financiamiento,
    solicitud_asesor,
    horario_contacto,
    nivel_lead,
    resumen,
):
    try:
        ahora = datetime.now()

        datos = {
            "fecha": ahora.strftime("%d/%m/%Y"),
            "hora": ahora.strftime("%H:%M:%S"),
            "nombre": nombre,
            "whatsapp": whatsapp,
            "id_whatsapp": id_whatsapp,
            "username": username,
            "motivo_compra": motivo_compra,
            "numero_personas": numero_personas,
            "inicial": inicial,
            "bono": bono,
            "ubicacion": ubicacion,
            "financiamiento": financiamiento,
            "solicitud_asesor": solicitud_asesor,
            "horario_contacto": horario_contacto,
            "nivel_lead": nivel_lead,
            "resumen": resumen,
        }

        respuesta = requests.post(
            GOOGLE_SHEETS_URL,
            json=datos,
            timeout=(5, 15),
        )

        print("GOOGLE SHEETS STATUS:", respuesta.status_code, flush=True)
        print("GOOGLE SHEETS RESPUESTA:", respuesta.text, flush=True)

        return respuesta.status_code == 200

    except Exception as error:
        print("ERROR GUARDANDO EN GOOGLE SHEETS:", error, flush=True)
        return False


# ==========================================
# CONVERTIR RESPUESTA DE IA EN CAMPOS
# ==========================================


def convertir_datos_cliente(texto):
    datos = {
        "Nombre": "No especificado",
        "Motivo de compra": "No especificado",
        "Número de personas": "No especificado",
        "Inicial": "No especificado",
        "Bono": "No especificado",
        "Ubicación": "No especificado",
        "Financiamiento": "No especificado",
        "Solicitud de asesor": "No especificado",
        "Horario de contacto": "No especificado",
        "Nivel del lead": "No especificado",
        "Resumen": "No especificado",
    }

    for linea in texto.splitlines():
        if ":" not in linea:
            continue

        clave, valor = linea.split(":", 1)
        clave = clave.strip()
        valor = valor.strip()

        if clave in datos:
            datos[clave] = valor

    return datos


# ==========================================
# EXTRAER DATOS DE TODA LA CONVERSACIÓN
# ==========================================

def extraer_texto_respuesta_openai(datos):
    partes = []
    for item in datos.get("output", []):
        if item.get("type") != "message":
            continue
        for contenido in item.get("content", []):
            if contenido.get("type") == "output_text":
                texto = contenido.get("text", "")
                if texto:
                    partes.append(texto)
    return "\n".join(partes).strip()


def llamar_openai_responses(instructions, input_text, max_output_tokens=220):
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY no está configurada en Render")

    url = "https://api.openai.com/v1/responses"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "Connection": "close",
    }
    payload = {
        "model": "gpt-5.6-luna",
        "reasoning": {"effort": "none"},
        "instructions": instructions,
        "input": input_text,
        "max_output_tokens": max_output_tokens,
    }

    session = requests.Session()
    session.trust_env = False
    try:
        print("OPENAI HTTP: iniciando solicitud")
        respuesta = session.post(
            url,
            headers=headers,
            json=payload,
            timeout=(5, 15),
        )
        print("OPENAI HTTP STATUS:", respuesta.status_code)
    except requests.Timeout as error:
        raise RuntimeError("OpenAI superó el tiempo máximo de espera (15 s)") from error
    except requests.RequestException as error:
        raise RuntimeError(f"No se pudo conectar con OpenAI: {error}") from error
    finally:
        session.close()

    if respuesta.status_code != 200:
        detalle = respuesta.text[:1200]
        raise RuntimeError(
            f"OpenAI devolvió HTTP {respuesta.status_code}: {detalle}"
        )

    datos = respuesta.json()
    texto = extraer_texto_respuesta_openai(datos)
    if not texto:
        raise RuntimeError(
            "OpenAI respondió correctamente, pero no devolvió texto utilizable"
        )
    return texto

def extraer_datos_cliente(numero):
    historial = historiales.get(numero, [])

    texto_historial = ""
    for mensaje in historial:
        texto_historial += (
            f"{mensaje['role']}: "
            f"{mensaje['content']}\n"
        )

    instructions = """
Analiza la conversación de un posible cliente interesado en Puerto Real.

Devuelve SOLO estos datos exactamente en este formato:

Nombre:
Motivo de compra:
Número de personas:
Inicial:
Bono:
Ubicación:
Financiamiento:
Solicitud de asesor:
Horario de contacto:
Nivel del lead:
Resumen:

REGLAS:
- No inventes información.
- Si no conoces un dato, escribe: No especificado
- Nivel del lead solo puede ser: FRÍO, TIBIO o CALIENTE.
- El resumen debe ser breve y útil para el equipo comercial.
- No confundas información proporcionada por el asistente con información confirmada por el cliente.
"""

    return llamar_openai_responses(
        instructions=instructions,
        input_text=texto_historial,
        max_output_tokens=450,
    )


# ==========================================
# VERIFICACIÓN DEL WEBHOOK
# ==========================================


@app.route("/webhook", methods=["GET"])
def verificar_webhook():
    modo = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    desafio = request.args.get("hub.challenge")

    print()
    print("===================================")
    print("VERIFICACIÓN DE WEBHOOK")
    print("===================================")

    if modo == "subscribe" and token == TOKEN_VERIFICACION:
        print("Webhook verificado correctamente")
        return desafio, 200

    print("Token de verificación incorrecto")
    return "Token incorrecto", 403


# ==========================================
# RECIBIR MENSAJES DE WHATSAPP
# ==========================================


@app.route("/webhook", methods=["POST"])
def recibir_mensaje():
    datos = request.get_json(silent=True) or {}

    try:
        value = datos["entry"][0]["changes"][0]["value"]
    except (KeyError, IndexError, TypeError):
        return "EVENT_RECEIVED", 200

    # Meta envía eventos de enviado, entregado, leído, etc.
    # No son mensajes del cliente y se ignoran silenciosamente.
    if "statuses" in value:
        print("ESTADO DE WHATSAPP:", value["statuses"], flush=True)
        return "EVENT_RECEIVED", 200

    mensajes = value.get("messages")
    if not mensajes:
        return "EVENT_RECEIVED", 200

    mensaje = mensajes[0]

    # Por ahora procesamos únicamente mensajes de texto.
    if mensaje.get("type") != "text":
        return "EVENT_RECEIVED", 200

    texto = mensaje.get("text", {}).get("body", "").strip()
    if not texto:
        return "EVENT_RECEIVED", 200

    contacto = value.get("contacts", [{}])[0]

    numero = (
        contacto.get("wa_id")
        or contacto.get("phone_number")
        or mensaje.get("from")
        or mensaje.get("from_user_id")
        or contacto.get("user_id")
    )

    if not numero:
        return "EVENT_RECEIVED", 200

    numero = str(numero)
    mensaje_id = mensaje.get("id")
    id_whatsapp = contacto.get("user_id", numero)

    perfil = contacto.get("profile", {})
    username = perfil.get("username", "No especificado")

    # Se registra antes de trabajar para bloquear reintentos de Meta.
    if not registrar_mensaje_procesado(mensaje_id):
        print("Mensaje duplicado ignorado:", mensaje_id)
        return "EVENT_RECEIVED", 200

    ahora = time.time()
    texto_normalizado = texto.lower()
    ultimo = ultimos_mensajes_por_cliente.get(numero)

    if ultimo:
        mismo_texto = ultimo["texto"] == texto_normalizado
        dentro_ventana = (ahora - ultimo["tiempo"]) <= VENTANA_DUPLICADO_SEGUNDOS

        if mismo_texto and dentro_ventana:
            print("Duplicado reciente ignorado:", numero, texto)
            return "EVENT_RECEIVED", 200

    ultimos_mensajes_por_cliente[numero] = {
        "texto": texto_normalizado,
        "tiempo": ahora,
    }

    print()
    print("===================================")
    print("MENSAJE RECIBIDO")
    print("===================================")
    print("Número:", numero)
    print("Mensaje:", texto)

    ultimo_mensaje_por_numero[numero] = mensaje_id

    print("INICIANDO PROCESAMIENTO DEL MENSAJE...")

    hilo = threading.Thread(
        target=procesar_mensaje_cliente,
        args=(
            numero,
            texto,
            mensaje_id,
            id_whatsapp,
            username,
        ),
        daemon=True,
    )
    hilo.start()

    print("HILO DE PROCESAMIENTO INICIADO")

    return "EVENT_RECEIVED", 200


# ==========================================
# PROCESAR MENSAJE DEL CLIENTE
# ==========================================

def enviar_imagen_whatsapp(numero, ruta_imagen, descripcion):
    token_whatsapp_actual = os.getenv("WHATSAPP_TOKEN")

    if not token_whatsapp_actual:
        print("WHATSAPP_TOKEN no está configurado.", flush=True)
        return False

    print("PREPARANDO IMAGEN:", ruta_imagen, flush=True)

    # Detectar tipo de imagen
    if ruta_imagen.lower().endswith(".png"):
        mime_type = "image/png"

    elif ruta_imagen.lower().endswith(".jpg") or ruta_imagen.lower().endswith(".jpeg"):
        mime_type = "image/jpeg"

    else:
        print("FORMATO DE IMAGEN NO SOPORTADO", flush=True)
        return False

    # ==========================================
    # 1. SUBIR IMAGEN A META
    # ==========================================

    url_media = f"https://graph.facebook.com/v26.0/{PHONE_NUMBER_ID}/media"

    headers = {
        "Authorization": f"Bearer {token_whatsapp_actual}"
    }

    try:
        with open(ruta_imagen, "rb") as archivo:

            files = {
                "file": (
                    os.path.basename(ruta_imagen),
                    archivo,
                    mime_type
                )
            }

            data = {
                "messaging_product": "whatsapp"
            }

            subida = requests.post(
                url_media,
                headers=headers,
                files=files,
                data=data,
                timeout=(5, 20)
            )

        print("SUBIDA IMAGEN:", subida.status_code, flush=True)
        print(subida.text, flush=True)

        if subida.status_code != 200:
            return False

        media_id = subida.json().get("id")

        if not media_id:
            print("NO SE RECIBIÓ MEDIA ID", flush=True)
            return False

        # ==========================================
        # 2. ENVIAR IMAGEN POR WHATSAPP
        # ==========================================

        url_mensaje = (
            f"https://graph.facebook.com/v26.0/"
            f"{PHONE_NUMBER_ID}/messages"
        )

        headers_mensaje = {
            "Authorization": f"Bearer {token_whatsapp_actual}",
            "Content-Type": "application/json",
        }

        payload = {
            "messaging_product": "whatsapp",
            "type": "image",
            "image": {
                "id": media_id,
                "caption": descripcion
            }
        }

        if "." in str(numero):
            payload["recipient"] = numero
        else:
            payload["to"] = numero

        envio = requests.post(
            url_mensaje,
            headers=headers_mensaje,
            json=payload,
            timeout=(5, 20)
        )

        print(
            "RESPUESTA IMAGEN WHATSAPP:",
            envio.status_code,
            flush=True
        )

        print(envio.text, flush=True)

        return envio.status_code == 200

    except FileNotFoundError:
        print(
            "ERROR: No se encontró la imagen:",
            ruta_imagen,
            flush=True
        )
        return False

    except Exception as error:
        print(
            "ERROR ENVIANDO IMAGEN:",
            error,
            flush=True
        )
        return False


def enviar_texto_whatsapp(numero, texto_respuesta):
    print("PREPARANDO ENVÍO A WHATSAPP...", flush=True)

    token_whatsapp_actual = os.getenv("WHATSAPP_TOKEN")

    if not token_whatsapp_actual:
        print("ERROR: WHATSAPP_TOKEN no está configurado.", flush=True)
        return False

    print("WHATSAPP_TOKEN encontrado.", flush=True)

    url = f"https://graph.facebook.com/v26.0/{PHONE_NUMBER_ID}/messages"

    headers = {
        "Authorization": f"Bearer {token_whatsapp_actual}",
        "Content-Type": "application/json",
    }

    payload = {
        "messaging_product": "whatsapp",
        "type": "text",
        "text": {
            "body": texto_respuesta
        },
    }

    if "." in str(numero):
        payload["recipient"] = numero
    else:
        payload["to"] = numero

    print("ENVIANDO RESPUESTA A WHATSAPP...", flush=True)
    print("DESTINATARIO:", numero, flush=True)

    try:
        envio = requests.post(
            url,
            headers=headers,
            json=payload,
            timeout=(5, 10)
        )

        print(
            "RESPUESTA DE WHATSAPP:",
            envio.status_code,
            flush=True
        )

        print(
            "DETALLE WHATSAPP:",
            envio.text,
            flush=True
        )

        return envio.status_code == 200

    except requests.Timeout:
        print(
            "ERROR: WhatsApp tardó demasiado en responder.",
            flush=True
        )
        return False

    except requests.RequestException as error:
        print(
            "ERROR ENVIANDO A WHATSAPP:",
            error,
            flush=True
        )
        return False

def detectar_imagen_para_mensaje(texto):
    t = texto.lower().strip()

    # 1) VARIAS UBICACIONES / ZONAS / PROYECTOS
    palabras_ubicaciones = [
        "ubicaciones",
        "que ubicaciones",
        "qué ubicaciones",
        "tienen ubicaciones",
        "zonas",
        "lugares",
        "proyectos",
        "donde tienen proyectos",
        "dónde tienen proyectos",
        "en que zonas",
        "en qué zonas",
    ]

    if any(p in t for p in palabras_ubicaciones):
        return None

    # 2) UBICACIÓN DE PUERTO REAL
    palabras_ubicacion = [
        "ubicacion",
        "ubicación",
        "donde",
        "dónde",
        "donde queda",
        "dónde queda",
        "donde esta",
        "dónde está",
        "ubicado",
        "ubicada",
        "pimentel",
    ]

    if any(p in t for p in palabras_ubicacion):
        return (
            IMAGEN_UBICACION,
            "Ubicación del proyecto Puerto Real en Pimentel 📍",
        )

    # 3) AVANCE / CONSTRUCCIÓN
    palabras_avance = [
        "constru",
        "avance",
        "piloto",
        "obra",
    ]

    if any(p in t for p in palabras_avance):
        return None

    # 4) CASA / FOTO
    palabras_casa = [
        "casa",
        "vivienda",
        "fachada",
        "como es",
        "cómo es",
        "foto",
        "fotos",
        "imagen",
        "imagenes",
        "imágenes",
    ]

    if any(p in t for p in palabras_casa):
        return (
            IMAGEN_CASA,
            "Así es el diseño de la vivienda de Puerto Real 🏡",
        )

    return None


def procesar_mensaje_cliente(
    numero,
    texto,
    mensaje_id,
    id_whatsapp,
    username,
):
    print("ENTRÓ A procesar_mensaje_cliente")

    if numero not in historiales:
        historiales[numero] = []

    historiales[numero].append({
        "role": "user",
        "content": texto,
    })

    texto_conversacion = ""
    for mensaje_historial in historiales[numero]:
        texto_conversacion += (
            f"{mensaje_historial['role']}: "
            f"{mensaje_historial['content']}\n"
        )

    instructions = f"""
Eres el asesor virtual de WhatsApp de Puerto Real, proyecto inmobiliario ubicado en Pimentel, Chiclayo.

IDENTIFICADOR DEL CLIENTE:
{numero}

BASE DE CONOCIMIENTO:
{informacion_puerto_real}

OBJETIVO:
Conversar como un asesor humano por WhatsApp, resolver dudas de manera breve y natural y, cuando exista interés real, registrar los datos necesarios para que el equipo comercial continúe la atención.

FORMA DE HABLAR:
- Español natural de Perú.
- Cercano, amable y profesional.
- Respuestas cortas: normalmente entre 1 y 4 líneas.
- Ve directo al punto.
- No envíes listas largas salvo que el cliente las pida.
- Usa 1 o 2 emojis cuando encajen.
- No seas seco ni robótico.
- No repitas saludos si la conversación ya comenzó.
- Responde primero exactamente a lo que el cliente acaba de decir.
- Haz máximo UNA pregunta por mensaje.
- No es obligatorio terminar cada mensaje con una pregunta.
- No repitas información que ya explicaste.
- Recuerda lo dicho anteriormente en el historial.

NOMBRE DEL CLIENTE:
- Si el cliente dice su nombre, recuérdalo durante la conversación.
- Úsalo de manera natural de vez en cuando.
- Si todavía no sabes su nombre y solicita hablar con un asesor, pregunta únicamente su nombre.

WHATSAPP:
- El sistema obtiene automáticamente el número o identificador de WhatsApp.
- NUNCA le pidas al cliente su número de celular.

ASESOR HUMANO:
Si el cliente quiere hablar con un asesor, ser contactado, comprar, separar, cotizar, visitar o avanzar:
1. Reconoce su intención.
2. No repitas toda la información del proyecto.
3. Si no sabes su nombre, pregunta únicamente su nombre.
4. Si ya sabes su nombre pero no su horario, pregunta únicamente el horario.
5. Si ya tienes nombre y horario, indica únicamente que la solicitud quedó registrada para el equipo comercial.
- No digas que ya agendaste, coordinaste o hablaste con un asesor.

BONO TECHO PROPIO:
- El Bono Techo Propio indicado para Puerto Real es de S/ 62,700.
- No determines si una persona califica.
- No inventes requisitos.
- Si pregunta si califica, indica que debe ser evaluado por el equipo comercial.

INFORMACIÓN FINANCIERA CONFIRMADA:
- Inicial: S/ 5,000.
- Bono Techo Propio: S/ 62,700.
- Financiamiento: hasta 48 meses sin intereses.

UBICACIÓN:
- Puerto Real está ubicado en Pimentel, Chiclayo.
- Si el cliente pregunta dónde queda el proyecto, dónde está ubicado o pide la ubicación, responde de forma breve y clara:
  "Puerto Real está ubicado en Pimentel, Chiclayo 😊📍"
- Si pregunta por varias ubicaciones, zonas o proyectos, no inventes otras ubicaciones; ofrece que un asesor comercial lo detalle.

CONTEXTO DEL PROYECTO:
HECHOS CONFIRMADOS:
- Puerto Real ofrece viviendas de interés social.
- Son casas de 2 pisos.
- Trabaja con el Bono Techo Propio de S/ 62,700.
- La entrega está proyectada en 30 meses.
- Actualmente se está por iniciar la construcción de la casa piloto.

CONTEXTO INTERNO:
- Al trabajar con Techo Propio y procesos vinculados al Estado, la documentación, planificación y desarrollo son más rigurosos y toman más tiempo.
- El proyecto avanza por etapas antes de construir todas las viviendas.
- Existe un proceso relacionado con alcanzar un porcentaje de interesados y gestionar los códigos correspondientes de Techo Propio.
- La vigencia del bono considerada para este proceso es de 1 año.

REGLAS SOBRE ESTE CONTEXTO:
- Es información interna para comprender el proceso.
- No expliques porcentajes, códigos, trámites ni vigencia del bono salvo que sea necesario.
- No inventes porcentajes, fechas, códigos ni avances.
- No digas que todas las casas ya están construidas o en construcción.
- No asegures que una persona ya tiene aprobado el bono.
- Si preguntan si ya están construyendo, explica brevemente que están próximos a iniciar la casa piloto y que el proyecto avanzará por etapas.
- Si preguntan por qué son 30 meses, explica de forma sencilla que por tratarse de vivienda de interés social con Techo Propio, la documentación, planificación y desarrollo avanzan por etapas.

INFORMACIÓN GENERAL:
- Si el cliente dice "quiero información", "más información", "informes", "quiero saber más" o algo similar, NO envíes toda la información del proyecto de golpe.
- Responde breve y de forma natural.
- Ejemplo:
  "¡Claro! 😊 Te ayudo con Puerto Real 🏡 ¿Qué te gustaría conocer: ubicación, características de la casa o facilidades de compra?"
- Después responde únicamente sobre el tema que el cliente elija.

RESPUESTAS ESPECÍFICAS:
- Si el cliente pregunta por ubicación, responde solo la ubicación.
- Si pregunta por precio, inicial, bono, cuotas o financiamiento, responde solo la información económica.
- Si pregunta por la casa, fachada o cómo es la vivienda, responde solo sobre la vivienda.
- No mezcles ubicación, precio y características en una sola respuesta si no lo pidió.

IMÁGENES:
- El sistema SÍ puede enviar imágenes.
- Nunca digas "no puedo enviar fotos" o "no puedo enviar imágenes".
- Si el cliente pide una foto, imagen, fachada o pregunta cómo es la vivienda, responde brevemente:
  "¡Claro! 😊 Te muestro el diseño de la vivienda de Puerto Real 🏡"
- Después del texto, el sistema enviará automáticamente la imagen de la casa.
- Si el cliente pregunta por ubicación, dónde queda o dónde está ubicado Puerto Real, responde:
  "Puerto Real está ubicado en Pimentel, Chiclayo 😊📍"
- Después del texto, el sistema enviará automáticamente la imagen de Pimentel.
- No digas que la imagen ya demuestra avance de obra o construcción terminada.

PRECIO Y FACILIDADES:
- Si el cliente pregunta por "precio", "precios", "cuánto cuesta", "inicial", "bono", "cuotas" o "financiamiento", responde únicamente con la información económica confirmada.
- Puedes responder:
  "La inicial es de S/ 5,000 😊 Además, contamos con Bono Techo Propio de S/ 62,700 y financiamiento hasta 48 meses sin intereses."
- Si pregunta por un precio final exacto que no está confirmado en la base de conocimiento, indica brevemente que un asesor comercial puede confirmarlo.

MUY IMPORTANTE:
- NO inventes precios, disponibilidad, promociones, requisitos, fórmulas ni fechas.
- NO calcules cuotas sin información suficiente.
- Si falta información, dilo brevemente y ofrece que un asesor comercial lo confirme.
- Si agradece o dice ok, responde brevemente y no fuerces otra pregunta.
"""

    input_actual = f"""
HISTORIAL DE LA CONVERSACIÓN:
{texto_conversacion}

MENSAJE ACTUAL DEL CLIENTE:
{texto}

Responde específicamente al MENSAJE ACTUAL DEL CLIENTE teniendo en cuenta el historial.
"""

    try:
        texto_normalizado = texto.lower().strip()

        if texto_normalizado in [
            "hola",
            "holaa",
            "buenas",
            "buenos dias",
            "buenas tardes",
        ]:
            respuesta_texto = (
                "¡Hola! 😊 Soy el asesor virtual de Puerto Real 🏡 "
                "¿En qué puedo ayudarte?"
            )

        elif any(p in texto_normalizado for p in [
            "quiero informacion",
            "quiero información",
            "mas informacion",
            "más información",
            "quiero mas informacion",
            "quiero más información",
            "informes",
        ]):
            respuesta_texto = (
                "¡Claro! 😊 Te ayudo con Puerto Real 🏡 "
                "¿Qué te gustaría conocer: ubicación, características de la casa o facilidades de compra?"
            )

        elif any(p in texto_normalizado for p in [
            "ubicaciones",
            "que ubicaciones",
            "qué ubicaciones",
            "tienen ubicaciones",
            "zonas",
            "lugares",
            "donde tienen proyectos",
            "dónde tienen proyectos",
            "en que zonas",
            "en qué zonas",
        ]):
            respuesta_texto = (
                "Sí contamos con información sobre nuestras ubicaciones y proyectos 😊 "
                "Un asesor comercial puede detallártelas mejor. "
                "¿Me indicas tu nombre para registrar tu solicitud?"
            )  
                    
        elif any(p in texto_normalizado for p in [
            "donde queda",
            "dónde queda",
            "ubicacion",
            "ubicación",
            "donde esta ubicado",
            "dónde está ubicado",
            "donde está",
            "dónde está",
        ]):
            respuesta_texto = (
                "Puerto Real está ubicado en Pimentel, Chiclayo 😊📍"
            )

        elif any(p in texto_normalizado for p in [
            "foto",
            "fotos",
            "imagen",
            "imagenes",
            "imágenes",
            "fachada",
            "mostrar la casa",
            "ver la casa",
        ]):
            respuesta_texto = (
                "¡Claro! 😊 Te muestro el diseño de la vivienda de Puerto Real 🏡"
            )

                # CALIFICACIÓN AL BONO -> DERIVAR A ASESOR
        elif any(p in texto_normalizado for p in [
            "califico",
            "calificar",
            "califica",
            "puedo calificar",
            "puedo acceder al bono",
            "puedo obtener el bono",
            "si califico",
            "si puedo obtener el bono",
            "me pueden evaluar",
            "evaluar para el bono",
        ]):
            respuesta_texto = (
                "Para saber si calificas al Bono Techo Propio es necesario que un asesor comercial "
                "realice la evaluación 😊 ¿Me indicas tu nombre para registrar tu solicitud?"
            )

        elif any(p in texto_normalizado for p in [
            "de cuanto es el bono",
            "de cuánto es el bono",
            "cuanto es el bono",
            "cuánto es el bono",
            "monto del bono",
        ]):
            respuesta_texto = (
                "El Bono Techo Propio para Puerto Real es de S/ 62,700 😊"
            )

        # PRECIO / BONO / FINANCIAMIENTO
        elif any(p in texto_normalizado for p in [
            "precio",
            "cuanto cuesta",
            "cuánto cuesta",
            "inicial",
            "bono",
            "financiamiento",
        ]):
            respuesta_texto = (
                "La inicial es de S/ 5,000 😊 Además, contamos con Bono Techo Propio "
                "de S/ 62,700 y financiamiento hasta 48 meses sin intereses."
            )
        else:
            print("ENVIANDO MENSAJE A OPENAI...")

            respuesta_texto = llamar_openai_responses(
                instructions=instructions,
                input_text=input_actual,
                max_output_tokens=220,
            )

            print("OPENAI RESPONDIÓ")

    except Exception as error:
        print("ERROR DE OPENAI:", error)
        respuesta_texto = (
            "Disculpa, en este momento estoy teniendo una demora para responder. "
            "Puedes intentar nuevamente en unos segundos 😊"
        )

    historiales[numero].append({
        "role": "assistant",
        "content": respuesta_texto,
    })

    # Primero enviamos el texto para que la conversación se sienta natural.
    enviar_texto_whatsapp(numero, respuesta_texto)

    # Después, si corresponde, enviamos UNA de las dos imágenes.
    imagen = detectar_imagen_para_mensaje(texto)
    if imagen:
        url_imagen, descripcion = imagen
        enviar_imagen_whatsapp(numero, url_imagen, descripcion)

    # CRM: un error aquí nunca debe impedir responder al cliente.
    try:
        datos_extraidos = extraer_datos_cliente(numero)
        datos_cliente = convertir_datos_cliente(datos_extraidos)

        telefono_detectado = re.search(
            r"\b(?:51)?9\d{8}\b",
            texto.replace(" ", ""),
        )
        whatsapp_para_excel = numero

        if telefono_detectado:
            telefono_real = telefono_detectado.group()
            if telefono_real.startswith("51"):
                telefono_real = telefono_real[2:]
            whatsapp_para_excel = telefono_real

        guardar_cliente_excel(
            datos_cliente["Nombre"],
            whatsapp_para_excel,
            id_whatsapp,
            username,
            datos_cliente["Motivo de compra"],
            datos_cliente["Número de personas"],
            datos_cliente["Inicial"],
            datos_cliente["Bono"],
            datos_cliente["Ubicación"],
            datos_cliente["Financiamiento"],
            datos_cliente["Solicitud de asesor"],
            datos_cliente["Horario de contacto"],
            datos_cliente["Nivel del lead"],
            datos_cliente["Resumen"],
        )

        guardar_cliente_google_sheets(
    datos_cliente["Nombre"],
    whatsapp_para_excel,
    id_whatsapp,
    username,
    datos_cliente["Motivo de compra"],
    datos_cliente["Número de personas"],
    datos_cliente["Inicial"],
    datos_cliente["Bono"],
    datos_cliente["Ubicación"],
    datos_cliente["Financiamiento"],
    datos_cliente["Solicitud de asesor"],
    datos_cliente["Horario de contacto"],
    datos_cliente["Nivel del lead"],
    datos_cliente["Resumen"],
)
        print("CLIENTE PROCESADO EN CRM")
    except Exception as error:
        print("ERROR GUARDANDO CLIENTE:", error)

    print("RESPUESTA FINAL:", respuesta_texto)


# ==========================================
# TRABAJADOR DE LA COLA
# ==========================================


def trabajador_mensajes():
    while True:
        numero, texto, mensaje_id, id_whatsapp, username = cola_mensajes.get()

        try:
            procesar_mensaje_cliente(
                numero,
                texto,
                mensaje_id,
                id_whatsapp,
                username,
            )
        except Exception as error:
            print("ERROR PROCESANDO MENSAJE:", error)
        finally:
            cola_mensajes.task_done()


# ==========================================
# INICIAR TRABAJADOR
# ==========================================

preparar_excel()
cargar_mensajes_procesados()


# ==========================================
# INICIAR SERVIDOR LOCAL
# ==========================================

if __name__ == "__main__":

    port = int(os.environ.get("PORT", 5001))

    app.run(
        host="0.0.0.0",
        port=port,
        debug=False,
        threaded=True,
    )
