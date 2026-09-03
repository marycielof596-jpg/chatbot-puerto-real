from flask import Flask, request
from openpyxl import load_workbook, Workbook
from datetime import datetime
import requests
import os
import threading
import queue
import time
import re


# ==========================================
# CONFIGURACIÓN
# ==========================================

app = Flask(__name__)

PHONE_NUMBER_ID = "1248995224968397"

IMAGEN_UBICACION = "https://raw.githubusercontent.com/marycielof596-jpg/chatbot-puerto-real/main/PIMENTEL.png"

IMAGEN_CASA = "https://raw.githubusercontent.com/marycielof596-jpg/chatbot-puerto-real/main/CASA%20PUERTO%20REAL%202026.jpeg"

TOKEN_VERIFICACION = "puerto_real_2026"

ARCHIVO_EXCEL = "clientes_puerto_real.xlsx"


# ==========================================
# MEMORIA DE CONVERSACIONES
# ==========================================

historiales = {}


# ==========================================
# CONTROL DE DUPLICADOS Y COLA DE TRABAJO
# ==========================================

ARCHIVO_MENSAJES_PROCESADOS = "mensajes_procesados.txt"

mensajes_procesados = set()
lock_mensajes = threading.Lock()
cola_mensajes = queue.Queue()
ultimo_mensaje_por_numero = {}

ultimos_mensajes_por_cliente = {}
VENTANA_DUPLICADO_SEGUNDOS = 12


def cargar_mensajes_procesados():
    if not os.path.exists(ARCHIVO_MENSAJES_PROCESADOS):
        return

    try:
        with open(ARCHIVO_MENSAJES_PROCESADOS, "r", encoding="utf-8") as archivo:
            for linea in archivo:
                mensaje_id = linea.strip()
                if mensaje_id:
                    mensajes_procesados.add(mensaje_id)
    except Exception as error:
        print("No se pudo cargar mensajes procesados:", error)


def registrar_mensaje_procesado(mensaje_id):
    if not mensaje_id:
        return True

    with lock_mensajes:
        if mensaje_id in mensajes_procesados:
            return False

        mensajes_procesados.add(mensaje_id)

        try:
            with open(ARCHIVO_MENSAJES_PROCESADOS, "a", encoding="utf-8") as archivo:
                archivo.write(mensaje_id + "\n")
        except Exception as error:
            print("No se pudo guardar el ID del mensaje:", error)

    return True


# ==========================================
# INFORMACIÓN DE PUERTO REAL
# ==========================================

informacion_puerto_real = """
PROYECTO: PUERTO REAL

Ubicación:
Pimentel, Chiclayo.

Tipo:
Viviendas.

Características:
- 2 pisos.
- 3 habitaciones.
- 2 baños.

Condominio:
- Condominio privado.
- Cerco perimétrico de concreto.
- Incluye 4 parques.

Construcción:
- Sistema constructivo tradicional de ladrillo y cemento.

Áreas:
- Área total: 60 m².
- Área construida: 60 m².
- Primer piso: 33 m².
- Segundo piso: 27 m².

Facilidades económicas:
- Inicial: S/ 5,000.
- Bono de Techo Propio: S/ 62,700.
- Financiamiento: 48 meses sin intereses.

Servicios básicos:
- Luz.
- Agua.
- Desagüe.

Entrega:
- Se entrega en 30 meses.
"""


# ==========================================
# EXCEL
# ==========================================

ENCABEZADOS_EXCEL = [
    "Fecha",
    "Hora",
    "Nombre",
    "WhatsApp",
    "ID WhatsApp",
    "Username",
    "Motivo de compra",
    "Número de personas",
    "Inicial",
    "Bono",
    "Ubicación",
    "Financiamiento",
    "Solicitud de asesor",
    "Horario de contacto",
    "Nivel del lead",
    "Resumen",
]


def preparar_excel():
    if not os.path.exists(ARCHIVO_EXCEL):
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Clientes"
        hoja.append(ENCABEZADOS_EXCEL)
        libro.save(ARCHIVO_EXCEL)
        return

    # Asegura que la fila de encabezados tenga las 16 columnas actuales.
    try:
        libro = load_workbook(ARCHIVO_EXCEL)
        hoja = libro.active
        for columna, encabezado in enumerate(ENCABEZADOS_EXCEL, start=1):
            hoja.cell(row=1, column=columna).value = encabezado
        libro.save(ARCHIVO_EXCEL)
    except PermissionError:
        print("ERROR: CIERRA EL EXCEL ANTES DE INICIAR EL SERVIDOR.")


def dato_valido(valor):
    if valor is None:
        return False

    valor = str(valor).strip()

    if valor == "":
        return False

    if valor.lower() in [
        "no especificado",
        "no especificada",
        "desconocido",
        "desconocida",
    ]:
        return False

    return True


def guardar_cliente_excel(
    nombre,
    whatsapp,
    id_whatsapp,
    username,
    motivo_compra,
    numero_personas,
    inicial,
    bono,
    ubicacion,
    financiamiento,
    solicitud_asesor,
    horario_contacto,
    nivel_lead,
    resumen,
):
    preparar_excel()

    ahora = datetime.now()
    fecha = ahora.strftime("%d/%m/%Y")
    hora = ahora.strftime("%H:%M")

    try:
        libro = load_workbook(ARCHIVO_EXCEL)
    except PermissionError:
        print()
        print("ERROR: CIERRA EL EXCEL ANTES DE GUARDAR.")
        print()
        return

    hoja = libro.active
    fila_encontrada = None

    for fila in range(2, hoja.max_row + 1):
        whatsapp_excel = hoja.cell(row=fila, column=4).value
        id_excel = hoja.cell(row=fila, column=5).value

        if (
            str(id_excel).strip() == str(id_whatsapp).strip()
            and str(id_whatsapp).strip() not in ["", "None", "No especificado"]
        ):
            fila_encontrada = fila
            break

        if (
            str(whatsapp_excel).strip() == str(whatsapp).strip()
            and str(whatsapp).strip() not in ["", "None", "No especificado"]
        ):
            fila_encontrada = fila
            break

    nuevos_datos = {
        3: nombre,
        4: whatsapp,
        5: id_whatsapp,
        6: username,
        7: motivo_compra,
        8: numero_personas,
        9: inicial,
        10: bono,
        11: ubicacion,
        12: financiamiento,
        13: solicitud_asesor,
        14: horario_contacto,
        15: nivel_lead,
        16: resumen,
    }

    if fila_encontrada:
        hoja.cell(row=fila_encontrada, column=1).value = fecha
        hoja.cell(row=fila_encontrada, column=2).value = hora

        for columna, valor in nuevos_datos.items():
            if dato_valido(valor):
                hoja.cell(row=fila_encontrada, column=columna).value = valor

        print("CLIENTE ACTUALIZADO EN EXCEL")

    else:
        hoja.append([
            fecha,
            hora,
            nombre,
            whatsapp,
            id_whatsapp,
            username,
            motivo_compra,
            numero_personas,
            inicial,
            bono,
            ubicacion,
            financiamiento,
            solicitud_asesor,
            horario_contacto,
            nivel_lead,
            resumen,
        ])

        print("CLIENTE GUARDADO EN EXCEL")

    libro.save(ARCHIVO_EXCEL)


# ==========================================
# CONVERTIR RESPUESTA DE IA EN CAMPOS
# ==========================================


def convertir_datos_cliente(texto):
    datos = {
        "Nombre": "No especificado",
        "Motivo de compra": "No especificado",
        "Número de personas": "No especificado",
        "Inicial": "No especificado",
        "Bono": "No especificado",
        "Ubicación": "No especificado",
        "Financiamiento": "No especificado",
        "Solicitud de asesor": "No especificado",
        "Horario de contacto": "No especificado",
        "Nivel del lead": "No especificado",
        "Resumen": "No especificado",
    }

    for linea in texto.splitlines():
        if ":" not in linea:
            continue

        clave, valor = linea.split(":", 1)
        clave = clave.strip()
        valor = valor.strip()

        if clave in datos:
            datos[clave] = valor

    return datos


# ==========================================
# EXTRAER DATOS DE TODA LA CONVERSACIÓN
# ==========================================

def extraer_texto_respuesta_openai(datos):
    partes = []
    for item in datos.get("output", []):
        if item.get("type") != "message":
            continue
        for contenido in item.get("content", []):
            if contenido.get("type") == "output_text":
                texto = contenido.get("text", "")
                if texto:
                    partes.append(texto)
    return "\n".join(partes).strip()


def llamar_openai_responses(instructions, input_text, max_output_tokens=220):
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY no está configurada en Render")

    url = "https://api.openai.com/v1/responses"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": "gpt-5.6-luna",
        "reasoning": {"effort": "none"},
        "instructions": instructions,
        "input": input_text,
        "max_output_tokens": max_output_tokens,
    }

    try:
        respuesta = requests.post(
            url,
            headers=headers,
            json=payload,
            timeout=(10, 25),
        )
    except requests.RequestException as error:
        raise RuntimeError(f"No se pudo conectar con OpenAI: {error}") from error

    if respuesta.status_code != 200:
        detalle = respuesta.text[:1200]
        raise RuntimeError(
            f"OpenAI devolvió HTTP {respuesta.status_code}: {detalle}"
        )

    datos = respuesta.json()
    texto = extraer_texto_respuesta_openai(datos)
    if not texto:
        raise RuntimeError(
            "OpenAI respondió correctamente, pero no devolvió texto utilizable"
        )
    return texto


def extraer_datos_cliente(numero):
    historial = historiales.get(numero, [])

    texto_historial = ""
    for mensaje in historial:
        texto_historial += (
            f"{mensaje['role']}: "
            f"{mensaje['content']}\n"
        )

    instructions = """
Analiza la conversación de un posible cliente interesado en Puerto Real.

Devuelve SOLO estos datos exactamente en este formato:

Nombre:
Motivo de compra:
Número de personas:
Inicial:
Bono:
Ubicación:
Financiamiento:
Solicitud de asesor:
Horario de contacto:
Nivel del lead:
Resumen:

REGLAS:
- No inventes información.
- Si no conoces un dato, escribe: No especificado
- Nivel del lead solo puede ser: FRÍO, TIBIO o CALIENTE.
- El resumen debe ser breve y útil para el equipo comercial.
- No confundas información proporcionada por el asistente con información confirmada por el cliente.
"""

    return llamar_openai_responses(
        instructions=instructions,
        input_text=texto_historial,
        max_output_tokens=450,
    )


# ==========================================
# VERIFICACIÓN DEL WEBHOOK
# ==========================================


@app.route("/webhook", methods=["GET"])
def verificar_webhook():
    modo = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    desafio = request.args.get("hub.challenge")

    print()
    print("===================================")
    print("VERIFICACIÓN DE WEBHOOK")
    print("===================================")

    if modo == "subscribe" and token == TOKEN_VERIFICACION:
        print("Webhook verificado correctamente")
        return desafio, 200

    print("Token de verificación incorrecto")
    return "Token incorrecto", 403


# ==========================================
# RECIBIR MENSAJES DE WHATSAPP
# ==========================================


@app.route("/webhook", methods=["POST"])
def recibir_mensaje():
    datos = request.get_json(silent=True) or {}

    try:
        value = datos["entry"][0]["changes"][0]["value"]
    except (KeyError, IndexError, TypeError):
        return "EVENT_RECEIVED", 200

    # Meta envía eventos de enviado, entregado, leído, etc.
    # No son mensajes del cliente y se ignoran silenciosamente.
    if "statuses" in value:
        return "EVENT_RECEIVED", 200

    mensajes = value.get("messages")
    if not mensajes:
        return "EVENT_RECEIVED", 200

    mensaje = mensajes[0]

    # Por ahora procesamos únicamente mensajes de texto.
    if mensaje.get("type") != "text":
        return "EVENT_RECEIVED", 200

    texto = mensaje.get("text", {}).get("body", "").strip()
    if not texto:
        return "EVENT_RECEIVED", 200

    contacto = value.get("contacts", [{}])[0]

    numero = (
        contacto.get("wa_id")
        or contacto.get("phone_number")
        or mensaje.get("from")
        or mensaje.get("from_user_id")
        or contacto.get("user_id")
    )

    if not numero:
        return "EVENT_RECEIVED", 200

    numero = str(numero)
    mensaje_id = mensaje.get("id")
    id_whatsapp = contacto.get("user_id", numero)

    perfil = contacto.get("profile", {})
    username = perfil.get("username", "No especificado")

    # Se registra antes de trabajar para bloquear reintentos de Meta.
    if not registrar_mensaje_procesado(mensaje_id):
        print("Mensaje duplicado ignorado:", mensaje_id)
        return "EVENT_RECEIVED", 200

    ahora = time.time()
    texto_normalizado = texto.lower()
    ultimo = ultimos_mensajes_por_cliente.get(numero)

    if ultimo:
        mismo_texto = ultimo["texto"] == texto_normalizado
        dentro_ventana = (ahora - ultimo["tiempo"]) <= VENTANA_DUPLICADO_SEGUNDOS

        if mismo_texto and dentro_ventana:
            print("Duplicado reciente ignorado:", numero, texto)
            return "EVENT_RECEIVED", 200

    ultimos_mensajes_por_cliente[numero] = {
        "texto": texto_normalizado,
        "tiempo": ahora,
    }

    print()
    print("===================================")
    print("MENSAJE RECIBIDO")
    print("===================================")
    print("Número:", numero)
    print("Mensaje:", texto)

    ultimo_mensaje_por_numero[numero] = mensaje_id

    print("MENSAJE ENVIADO A LA COLA DE PROCESAMIENTO")

    cola_mensajes.put((
        numero,
        texto,
        mensaje_id,
        id_whatsapp,
        username,
    ))

    return "EVENT_RECEIVED", 200


# ==========================================
# PROCESAR MENSAJE DEL CLIENTE
# ==========================================

def enviar_imagen_whatsapp(numero, url_imagen, descripcion):
    token_whatsapp_actual = os.getenv("WHATSAPP_TOKEN")
    if not token_whatsapp_actual:
        print("WHATSAPP_TOKEN no está configurado.")
        return False

    url = f"https://graph.facebook.com/v26.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {token_whatsapp_actual}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "type": "image",
        "image": {
            "link": url_imagen,
            "caption": descripcion,
        },
    }

    if "." in str(numero):
        payload["recipient"] = numero
    else:
        payload["to"] = numero

    try:
        respuesta = requests.post(
            url, headers=headers, json=payload, timeout=(10, 20)
        )
        print("RESPUESTA IMAGEN WHATSAPP:", respuesta.status_code)
        print(respuesta.text)
        return respuesta.status_code == 200
    except requests.RequestException as error:
        print("ERROR ENVIANDO IMAGEN A WHATSAPP:", error)
        return False


def enviar_texto_whatsapp(numero, texto_respuesta):
    token_whatsapp_actual = os.getenv("WHATSAPP_TOKEN")
    if not token_whatsapp_actual:
        print("WHATSAPP_TOKEN no está configurado.")
        return False

    url = f"https://graph.facebook.com/v26.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {token_whatsapp_actual}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "type": "text",
        "text": {"body": texto_respuesta},
    }

    if "." in str(numero):
        payload["recipient"] = numero
    else:
        payload["to"] = numero

    try:
        envio = requests.post(
            url, headers=headers, json=payload, timeout=(10, 20)
        )
        print("RESPUESTA DE WHATSAPP:", envio.status_code)
        print(envio.text)
        return envio.status_code == 200
    except requests.RequestException as error:
        print("ERROR ENVIANDO A WHATSAPP:", error)
        return False


def detectar_imagen_para_mensaje(texto):
    t = texto.lower().strip()

    palabras_ubicacion = [
        "ubicacion", "ubicación", "donde queda", "dónde queda",
        "donde esta", "dónde está", "ubicado", "ubicada", "pimentel",
    ]
    if any(p in t for p in palabras_ubicacion):
        return (
            IMAGEN_UBICACION,
            "Puerto Real está ubicado en Pimentel, Chiclayo 📍",
        )

    # Si pregunta por avance/construcción, no mandamos el render para no dar
    # la impresión de que esa vivienda ya está construida.
    palabras_avance = ["constru", "avance", "piloto", "obra"]
    if any(p in t for p in palabras_avance):
        return None

    palabras_casa = [
        "casa", "vivienda", "fachada", "como es", "cómo es",
        "foto", "fotos", "imagen", "imagenes", "imágenes",
    ]
    if any(p in t for p in palabras_casa):
        return (
            IMAGEN_CASA,
            "Así es el diseño de la vivienda de Puerto Real 🏡",
        )

    return None


def procesar_mensaje_cliente(
    numero,
    texto,
    mensaje_id,
    id_whatsapp,
    username,
):
    print("ENTRÓ A procesar_mensaje_cliente")

    if numero not in historiales:
        historiales[numero] = []

    historiales[numero].append({
        "role": "user",
        "content": texto,
    })

    texto_conversacion = ""
    for mensaje_historial in historiales[numero]:
        texto_conversacion += (
            f"{mensaje_historial['role']}: "
            f"{mensaje_historial['content']}\n"
        )

    instructions = f"""
Eres el asesor virtual de WhatsApp de Puerto Real, proyecto inmobiliario ubicado en Pimentel, Chiclayo.

IDENTIFICADOR DEL CLIENTE:
{numero}

BASE DE CONOCIMIENTO:
{informacion_puerto_real}

OBJETIVO:
Conversar como un asesor humano por WhatsApp, resolver dudas de manera breve y natural y, cuando exista interés real, registrar los datos necesarios para que el equipo comercial continúe la atención.

FORMA DE HABLAR:
- Español natural de Perú.
- Cercano, amable y profesional.
- Respuestas cortas: normalmente entre 1 y 4 líneas.
- Ve directo al punto.
- No envíes listas largas salvo que el cliente las pida.
- Usa 1 o 2 emojis cuando encajen.
- No seas seco ni robótico.
- No repitas saludos si la conversación ya comenzó.
- Responde primero exactamente a lo que el cliente acaba de decir.
- Haz máximo UNA pregunta por mensaje.
- No es obligatorio terminar cada mensaje con una pregunta.
- No repitas información que ya explicaste.
- Recuerda lo dicho anteriormente en el historial.

NOMBRE DEL CLIENTE:
- Si el cliente dice su nombre, recuérdalo durante la conversación.
- Úsalo de manera natural de vez en cuando.
- Si todavía no sabes su nombre y solicita hablar con un asesor, pregunta únicamente su nombre.

WHATSAPP:
- El sistema obtiene automáticamente el número o identificador de WhatsApp.
- NUNCA le pidas al cliente su número de celular.

ASESOR HUMANO:
Si el cliente quiere hablar con un asesor, ser contactado, comprar, separar, cotizar, visitar o avanzar:
1. Reconoce su intención.
2. No repitas toda la información del proyecto.
3. Si no sabes su nombre, pregunta únicamente su nombre.
4. Si ya sabes su nombre pero no su horario, pregunta únicamente el horario.
5. Si ya tienes nombre y horario, indica únicamente que la solicitud quedó registrada para el equipo comercial.
- No digas que ya agendaste, coordinaste o hablaste con un asesor.

BONO TECHO PROPIO:
- El Bono Techo Propio indicado para Puerto Real es de S/ 62,700.
- No determines si una persona califica.
- No inventes requisitos.
- Si pregunta si califica, indica que debe ser evaluado por el equipo comercial.

INFORMACIÓN FINANCIERA CONFIRMADA:
- Inicial: S/ 5,000.
- Bono Techo Propio: S/ 62,700.
- Financiamiento: hasta 48 meses sin intereses.

UBICACIÓN:
- Puerto Real está ubicado en Pimentel, Chiclayo.
- Si pregunta por varias ubicaciones, zonas o proyectos, no inventes otras ubicaciones; ofrece que un asesor comercial lo detalle.

CONTEXTO DEL PROYECTO:
HECHOS CONFIRMADOS:
- Puerto Real ofrece viviendas de interés social.
- Son casas de 2 pisos.
- Trabaja con el Bono Techo Propio de S/ 62,700.
- La entrega está proyectada en 30 meses.
- Actualmente se está por iniciar la construcción de la casa piloto.

CONTEXTO INTERNO:
- Al trabajar con Techo Propio y procesos vinculados al Estado, la documentación, planificación y desarrollo son más rigurosos y toman más tiempo.
- El proyecto avanza por etapas antes de construir todas las viviendas.
- Existe un proceso relacionado con alcanzar un porcentaje de interesados y gestionar los códigos correspondientes de Techo Propio.
- La vigencia del bono considerada para este proceso es de 1 año.

REGLAS SOBRE ESTE CONTEXTO:
- Es información interna para comprender el proceso.
- No expliques porcentajes, códigos, trámites ni vigencia del bono salvo que sea necesario.
- No inventes porcentajes, fechas, códigos ni avances.
- No digas que todas las casas ya están construidas o en construcción.
- No asegures que una persona ya tiene aprobado el bono.
- Si preguntan si ya están construyendo, explica brevemente que están próximos a iniciar la casa piloto y que el proyecto avanzará por etapas.
- Si preguntan por qué son 30 meses, explica de forma sencilla que por tratarse de vivienda de interés social con Techo Propio, la documentación, planificación y desarrollo avanzan por etapas.

MUY IMPORTANTE:
- NO inventes precios, disponibilidad, promociones, requisitos, fórmulas ni fechas.
- NO calcules cuotas sin información suficiente.
- Si falta información, dilo brevemente y ofrece que un asesor comercial lo confirme.
- Si agradece o dice ok, responde brevemente y no fuerces otra pregunta.
"""

    input_actual = f"""
HISTORIAL DE LA CONVERSACIÓN:
{texto_conversacion}

MENSAJE ACTUAL DEL CLIENTE:
{texto}

Responde específicamente al MENSAJE ACTUAL DEL CLIENTE teniendo en cuenta el historial.
"""

    try:
        print("ENVIANDO MENSAJE A OPENAI...")
        respuesta_texto = llamar_openai_responses(
            instructions=instructions,
            input_text=input_actual,
            max_output_tokens=220,
        )
        print("OPENAI RESPONDIÓ")
    except Exception as error:
        print("ERROR DE OPENAI:", error)
        respuesta_texto = (
            "Disculpa, en este momento estoy teniendo una demora para responder. "
            "Puedes intentar nuevamente en unos segundos 😊"
        )

    historiales[numero].append({
        "role": "assistant",
        "content": respuesta_texto,
    })

    # Primero enviamos el texto para que la conversación se sienta natural.
    enviar_texto_whatsapp(numero, respuesta_texto)

    # Después, si corresponde, enviamos UNA de las dos imágenes.
    imagen = detectar_imagen_para_mensaje(texto)
    if imagen:
        url_imagen, descripcion = imagen
        enviar_imagen_whatsapp(numero, url_imagen, descripcion)

    # CRM: un error aquí nunca debe impedir responder al cliente.
    try:
        datos_extraidos = extraer_datos_cliente(numero)
        datos_cliente = convertir_datos_cliente(datos_extraidos)

        telefono_detectado = re.search(
            r"\b(?:51)?9\d{8}\b",
            texto.replace(" ", ""),
        )
        whatsapp_para_excel = numero

        if telefono_detectado:
            telefono_real = telefono_detectado.group()
            if telefono_real.startswith("51"):
                telefono_real = telefono_real[2:]
            whatsapp_para_excel = telefono_real

        guardar_cliente_excel(
            datos_cliente["Nombre"],
            whatsapp_para_excel,
            id_whatsapp,
            username,
            datos_cliente["Motivo de compra"],
            datos_cliente["Número de personas"],
            datos_cliente["Inicial"],
            datos_cliente["Bono"],
            datos_cliente["Ubicación"],
            datos_cliente["Financiamiento"],
            datos_cliente["Solicitud de asesor"],
            datos_cliente["Horario de contacto"],
            datos_cliente["Nivel del lead"],
            datos_cliente["Resumen"],
        )
        print("CLIENTE PROCESADO EN CRM")
    except Exception as error:
        print("ERROR GUARDANDO CLIENTE:", error)

    print("RESPUESTA FINAL:", respuesta_texto)


# ==========================================
# TRABAJADOR DE LA COLA
# ==========================================


def trabajador_mensajes():
    while True:
        numero, texto, mensaje_id, id_whatsapp, username = cola_mensajes.get()

        try:
            procesar_mensaje_cliente(
                numero,
                texto,
                mensaje_id,
                id_whatsapp,
                username,
            )
        except Exception as error:
            print("ERROR PROCESANDO MENSAJE:", error)
        finally:
            cola_mensajes.task_done()


# ==========================================
# INICIAR TRABAJADOR
# ==========================================

preparar_excel()
cargar_mensajes_procesados()

hilo_trabajador = threading.Thread(
    target=trabajador_mensajes,
    daemon=True,
)

hilo_trabajador.start()


# ==========================================
# INICIAR SERVIDOR LOCAL
# ==========================================

if __name__ == "__main__":

    port = int(os.environ.get("PORT", 5001))

    app.run(
        host="0.0.0.0",
        port=port,
        debug=False,
        threaded=True,
    )
